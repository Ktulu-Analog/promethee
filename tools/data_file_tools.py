# ============================================================================
# Prométhée — Assistant IA avancé
# ============================================================================
# Auteur  : Pierre COUGET ktulu.analog@gmail.com
# Licence : GNU Affero General Public License v3.0 (AGPL-3.0)
#           https://www.gnu.org/licenses/agpl-3.0.html
# Année   : 2026
# ----------------------------------------------------------------------------
# Ce fichier fait partie du projet Prométhée.
# Vous pouvez le redistribuer et/ou le modifier selon les termes de la
# licence AGPL-3.0 publiée par la Free Software Foundation.
# ============================================================================

"""
tools/data_file_tools.py — Outils CSV / Excel avancés
=======================================================

Outils exposés (27) :

  Lecture (3) :
    - df_read           : charge un CSV ou Excel (avec option nrows et limite taille)
    - df_read_sheets    : charge toutes les feuilles d'un Excel en datasets distincts
    - df_list           : liste les datasets chargés en session

  Exploration (3) :
    - df_head           : premières/dernières lignes
    - df_info           : statistiques descriptives complètes + détection types
    - df_value_counts   : fréquence des valeurs d'une colonne

  Analyse (5) :
    - df_groupby        : agrégation GROUP BY avec support périodes temporelles
    - df_resample       : agrégation par période temporelle (mois, trimestre, année…)
    - df_correlate      : matrice de corrélation
    - df_outliers       : détection valeurs aberrantes (IQR ou z-score)
    - df_compare        : diff entre deux versions d'un dataset (nouveaux/supprimés/modifiés)

  Transformation (9) :
    - df_query          : filtre / sélectionne / trie
    - df_pivot          : tableau croisé dynamique
    - df_merge          : JOIN entre deux datasets
    - df_concat         : UNION verticale
    - df_clean          : nettoyage en une passe
    - df_cast           : conversion de types
    - df_apply          : colonnes calculées (moteur d'expressions sécurisé)
    - df_rename         : renommage de colonnes
    - df_duplicates     : détection doublons (exact + fuzzy)

  Échantillonnage (1) :
    - df_sample         : tirage aléatoire ou stratifié

  Conformité (1) :
    - df_anonymize      : anonymisation / pseudonymisation RGPD

  Écriture (3) :
    - df_write          : export CSV ou Excel formaté
    - df_write_excel    : export Excel multi-feuilles avec mise en forme complète
    - df_drop           : supprime un dataset de la mémoire session

Correctifs v2 :
  - Isolation _DATASETS par utilisateur (clé = user_id)
  - Limite de taille à la lecture (MAX_FILE_SIZE_MB, configurable)
  - Option nrows dans df_read pour exploration rapide de grands fichiers
  - df_apply : eval remplacé par un moteur d'expressions restreint (pas de __class__, etc.)
  - df_groupby : support groupby par période temporelle (M, Q, Y…)
  - df_write : formatage Excel (en-têtes gras, largeurs auto, freeze pane)
  - df_list : affiche la taille RAM de chaque dataset
"""

import hashlib
import io as _io
import json
import math
import re
import secrets
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import numpy as np
import pandas as pd

from core.tools_engine import tool, set_current_family, _TOOL_ICONS
from core.request_context import get_user_config

set_current_family("data_file_tools", "Fichiers de données", "📊")

_TOOL_ICONS.update({
    "df_read":         "📂",
    "df_read_sheets":  "📑",
    "df_list":         "📋",
    "df_head":         "👁️",
    "df_info":         "📊",
    "df_value_counts": "🔢",
    "df_groupby":      "📦",
    "df_resample":     "📅",
    "df_correlate":    "🔗",
    "df_outliers":     "⚠️",
    "df_compare":      "🔀",
    "df_query":        "🔍",
    "df_pivot":        "🔄",
    "df_merge":        "🔗",
    "df_concat":       "📎",
    "df_clean":        "🧹",
    "df_cast":         "🔁",
    "df_apply":        "⚙️",
    "df_rename":       "✏️",
    "df_duplicates":   "👥",
    "df_sample":       "🎲",
    "df_anonymize":    "🔒",
    "df_write":        "💾",
    "df_write_excel":  "📗",
    "df_drop":         "🗑️",
})

# ── Registre de datasets isolé par utilisateur ────────────────────────────────
# { user_id: { nom_dataset: { "df": DataFrame, "source": str, "loaded_at": str } } }
_DATASETS_BY_USER: dict[str, dict[str, dict]] = {}

_MAX_ROWS_DISPLAY  = 200
_MAX_COLS_DISPLAY  = 50
_MAX_FILE_SIZE_MB  = 100   # limite lecture VFS


# ── Helpers ───────────────────────────────────────────────────────────────────

def _uid() -> str:
    ucfg = get_user_config()
    return ucfg.user_id if ucfg else "default"


def _user_datasets() -> dict:
    uid = _uid()
    if uid not in _DATASETS_BY_USER:
        _DATASETS_BY_USER[uid] = {}
    return _DATASETS_BY_USER[uid]


def _get_df(nom: str) -> pd.DataFrame:
    ds = _user_datasets()
    if nom not in ds:
        noms = list(ds.keys())
        hint = f" Datasets disponibles : {noms}." if noms else " Aucun dataset chargé."
        raise KeyError(
            f"Dataset '{nom}' introuvable.{hint} "
            "Utilisez df_read pour charger un fichier."
        )
    return ds[nom]["df"]


def _store(nom: str, df: pd.DataFrame, source: str) -> None:
    _user_datasets()[nom] = {
        "df":        df,
        "source":    source,
        "loaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _safe(val: Any) -> Any:
    if val is None:
        return None
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return None
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.isoformat()
    if isinstance(val, pd.NaT.__class__):
        return None
    if isinstance(val, np.integer):
        return int(val)
    if isinstance(val, np.floating):
        v = float(val)
        return None if (math.isnan(v) or math.isinf(v)) else v
    if isinstance(val, np.bool_):
        return bool(val)
    if isinstance(val, np.ndarray):
        return val.tolist()
    if isinstance(val, bytes):
        return f"<bytes {len(val)}>"
    return val


def _df_to_records(df: pd.DataFrame, max_rows: int = _MAX_ROWS_DISPLAY) -> tuple[list, bool]:
    truncated = len(df) > max_rows
    subset = df.head(max_rows)
    records = []
    for row in subset.itertuples(index=False):
        records.append({col: _safe(val) for col, val in zip(subset.columns, row)})
    return records, truncated


def _detect_encoding(raw: bytes) -> str:
    try:
        if raw[:4].startswith(b"\xef\xbb\xbf"):
            return "utf-8-sig"
        raw[:8192].decode("utf-8")
        return "utf-8"
    except Exception:
        return "latin-1"


def _infer_separator(raw: bytes, encoding: str) -> str:
    try:
        first_line = raw.decode(encoding, errors="replace").splitlines()[0]
        counts = {sep: first_line.count(sep) for sep in (",", ";", "\t", "|")}
        return max(counts, key=counts.get)
    except Exception:
        return ","


def _ram_size(df: pd.DataFrame) -> str:
    b = df.memory_usage(deep=True).sum()
    if b < 1024:
        return f"{b} o"
    if b < 1_048_576:
        return f"{b/1024:.1f} Ko"
    return f"{b/1_048_576:.1f} Mo"


# ── Moteur d'expressions sécurisé pour df_apply ───────────────────────────────
# Remplace eval(formule, {"__builtins__": {}}, ...) qui est contournable.
# Vérifie l'AST avant toute évaluation pour bloquer les accès dangereux.

def _safe_eval(formule: str, context: dict) -> Any:
    import ast

    _BLACKLIST_ATTRS = {
        "__class__", "__bases__", "__subclasses__", "__globals__",
        "__builtins__", "__import__", "__loader__", "__spec__",
        "__code__", "__func__", "__self__", "__dict__", "__module__",
        "__mro__", "mro", "__init_subclass__",
    }

    class _ASTChecker(ast.NodeVisitor):
        def visit_Attribute(self, node):
            if isinstance(node.attr, str) and node.attr in _BLACKLIST_ATTRS:
                raise ValueError(f"Attribut interdit dans l'expression : '{node.attr}'")
            self.generic_visit(node)

        def visit_Call(self, node):
            if isinstance(node.func, ast.Name):
                if node.func.id in ("eval", "exec", "compile", "open", "__import__",
                                    "getattr", "setattr", "delattr", "vars", "dir",
                                    "globals", "locals", "breakpoint"):
                    raise ValueError(f"Fonction interdite : '{node.func.id}'")
            self.generic_visit(node)

    try:
        tree = ast.parse(formule, mode="eval")
    except SyntaxError as e:
        raise ValueError(f"Syntaxe invalide : {e}")

    _ASTChecker().visit(tree)

    safe_builtins = {
        "abs": abs, "round": round, "min": min, "max": max,
        "int": int, "float": float, "str": str, "bool": bool,
        "len": len, "sum": sum, "list": list, "dict": dict,
        "True": True, "False": False, "None": None,
    }
    return eval(  # noqa: S307 — AST validé ci-dessus
        compile(tree, "<expr>", "eval"),
        {"__builtins__": safe_builtins},
        context,
    )


# ══════════════════════════════════════════════════════════════════════════════
# LECTURE
# ══════════════════════════════════════════════════════════════════════════════

@tool(
    name="df_read",
    description=(
        "Charge un fichier CSV ou Excel (.xlsx, .xls, .ods) depuis le VFS en mémoire "
        "sous un nom court pour la session. "
        "Détecte automatiquement l'encodage et le séparateur CSV. "
        "Pour les grands fichiers, utiliser 'nrows' pour charger un sous-ensemble. "
        "Le dataset reste disponible pour toute la session sous le nom donné."
    ),
    parameters={
        "type": "object",
        "properties": {
            "chemin": {
                "type": "string",
                "description": "Chemin VFS du fichier (ex: /documents/agents.xlsx).",
            },
            "nom": {
                "type": "string",
                "description": "Nom court pour référencer le dataset (défaut: nom du fichier sans extension).",
            },
            "feuille": {
                "type": "string",
                "description": "Nom ou index (0-basé) de la feuille Excel (défaut: première feuille).",
            },
            "separateur": {
                "type": "string",
                "description": "Séparateur CSV (défaut: détection automatique parmi , ; \\t |).",
            },
            "encodage": {
                "type": "string",
                "description": "Encodage du fichier CSV (défaut: détection automatique).",
            },
            "lignes_header": {
                "type": "integer",
                "description": "Numéro de la ligne d'en-têtes 0-basé (défaut: 0).",
            },
            "ignorer_lignes": {
                "type": "integer",
                "description": "Nombre de lignes à sauter en début de fichier (défaut: 0).",
            },
            "nrows": {
                "type": "integer",
                "description": "Nombre maximum de lignes à charger (défaut: toutes). Utile pour explorer un grand fichier.",
            },
        },
        "required": ["chemin"],
    },
)
def df_read(
    chemin: str,
    nom: Optional[str] = None,
    feuille: Optional[str] = None,
    separateur: Optional[str] = None,
    encodage: Optional[str] = None,
    lignes_header: int = 0,
    ignorer_lignes: int = 0,
    nrows: Optional[int] = None,
) -> dict:
    from core.virtual_fs import get_vfs

    vfs = get_vfs()
    if not vfs.exists(chemin) or not vfs.is_file(chemin):
        return {"status": "error", "error": f"Fichier introuvable dans le VFS : {chemin}"}

    # Vérification taille avant lecture complète
    try:
        info = vfs.info(chemin)
        size_bytes = info.get("size", 0) if info else 0
        size_mb = size_bytes / 1_048_576
        if size_mb > _MAX_FILE_SIZE_MB:
            return {
                "status": "error",
                "error": (
                    f"Fichier trop volumineux : {size_mb:.1f} Mo "
                    f"(limite : {_MAX_FILE_SIZE_MB} Mo). "
                    "Utilisez 'nrows' pour charger un sous-ensemble, "
                    "ou découpez le fichier en amont."
                ),
            }
    except Exception:
        pass

    raw = vfs.read_bytes(chemin)
    dataset_name = nom or Path(chemin).stem
    ext = Path(chemin).suffix.lower()

    try:
        t0 = time.perf_counter()

        if ext in (".xlsx", ".xls", ".ods", ".xlsm"):
            engine = "openpyxl" if ext != ".xls" else "xlrd"
            xl = pd.ExcelFile(_io.BytesIO(raw), engine=engine)
            sheet_names = xl.sheet_names

            sheet = feuille
            if sheet is None:
                sheet = sheet_names[0]
            elif isinstance(sheet, str) and sheet.isdigit():
                sheet = sheet_names[int(sheet)]

            df = pd.read_excel(
                xl, sheet_name=sheet,
                header=lignes_header,
                skiprows=range(1, ignorer_lignes + 1) if ignorer_lignes > 0 else None,
                nrows=nrows,
            )
            source_info = f"{Path(chemin).name} / feuille '{sheet}'"
            extra = {"feuilles_disponibles": sheet_names, "feuille_chargee": sheet}

        elif ext in (".csv", ".tsv", ".txt"):
            enc = encodage or _detect_encoding(raw)
            sep = separateur or _infer_separator(raw, enc)

            df = pd.read_csv(
                _io.BytesIO(raw),
                sep=sep,
                encoding=enc,
                header=lignes_header,
                skiprows=range(1, ignorer_lignes + 1) if ignorer_lignes > 0 else None,
                nrows=nrows,
                low_memory=False,
                na_values=["", "NA", "N/A", "NULL", "null", "None", "#N/A"],
            )
            source_info = Path(chemin).name
            extra = {"separateur_detecte": sep, "encodage_detecte": enc}

        else:
            return {
                "status": "error",
                "error": f"Format non supporté : '{ext}'. Formats acceptés : .csv .tsv .xlsx .xls .ods",
            }

        elapsed_ms = round((time.perf_counter() - t0) * 1000, 1)
        df.columns = [str(c).strip() for c in df.columns]

        _store(dataset_name, df, source_info)

        # Détection automatique des colonnes date mal typées (object mais parseable)
        date_hints = []
        for col in df.select_dtypes(include="object").columns:
            sample = df[col].dropna().head(20)
            if len(sample) >= 3:
                try:
                    pd.to_datetime(sample, infer_datetime_format=True, dayfirst=True)
                    date_hints.append(col)
                except Exception:
                    pass

        result = {
            "status":      "success",
            "nom":         dataset_name,
            "source":      source_info,
            "nb_lignes":   len(df),
            "nb_colonnes": len(df.columns),
            "colonnes":    list(df.columns),
            "types":       {col: str(dtype) for col, dtype in df.dtypes.items()},
            "ram":         _ram_size(df),
            "duree_ms":    elapsed_ms,
            "message":     f"Dataset '{dataset_name}' chargé : {len(df)} lignes × {len(df.columns)} colonnes.",
            **extra,
        }
        if date_hints:
            result["colonnes_date_potentielles"] = date_hints
            result["conseil"] = (
                f"Les colonnes {date_hints} semblent contenir des dates mais sont typées 'object'. "
                "Utilisez df_cast pour les convertir en 'datetime'."
            )
        if nrows and len(df) == nrows:
            result["avertissement"] = f"Chargement limité à {nrows} lignes (paramètre nrows). Le fichier peut contenir plus de données."
        return result

    except Exception as e:
        return {"status": "error", "error": f"Erreur lecture : {e}"}


@tool(
    name="df_read_sheets",
    description=(
        "Charge toutes les feuilles d'un fichier Excel en datasets distincts en une seule opération. "
        "Chaque feuille devient un dataset nommé '<prefixe>_<nom_feuille>'. "
        "Utile pour les exports SIRH ou comptables avec plusieurs onglets."
    ),
    parameters={
        "type": "object",
        "properties": {
            "chemin": {
                "type": "string",
                "description": "Chemin VFS du fichier Excel.",
            },
            "prefixe": {
                "type": "string",
                "description": "Préfixe pour les noms de datasets (défaut: nom du fichier sans extension).",
            },
            "feuilles": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Liste des feuilles à charger (défaut: toutes).",
            },
            "nrows": {
                "type": "integer",
                "description": "Nombre max de lignes par feuille (défaut: toutes).",
            },
        },
        "required": ["chemin"],
    },
)
def df_read_sheets(
    chemin: str,
    prefixe: Optional[str] = None,
    feuilles: Optional[list] = None,
    nrows: Optional[int] = None,
) -> dict:
    from core.virtual_fs import get_vfs

    vfs = get_vfs()
    if not vfs.exists(chemin) or not vfs.is_file(chemin):
        return {"status": "error", "error": f"Fichier introuvable : {chemin}"}

    ext = Path(chemin).suffix.lower()
    if ext not in (".xlsx", ".xlsm", ".xls", ".ods"):
        return {"status": "error", "error": "df_read_sheets ne supporte que les fichiers Excel."}

    try:
        raw = vfs.read_bytes(chemin)
        engine = "openpyxl" if ext != ".xls" else "xlrd"
        xl = pd.ExcelFile(_io.BytesIO(raw), engine=engine)
        pfx = prefixe or Path(chemin).stem

        sheets_to_load = feuilles or xl.sheet_names
        chargees = []

        for sheet in sheets_to_load:
            if sheet not in xl.sheet_names:
                continue
            df = pd.read_excel(xl, sheet_name=sheet, nrows=nrows)
            df.columns = [str(c).strip() for c in df.columns]
            ds_name = f"{pfx}_{re.sub(r'[^a-zA-Z0-9_]', '_', sheet)}"
            _store(ds_name, df, f"{Path(chemin).name} / {sheet}")
            chargees.append({
                "feuille":     sheet,
                "nom_dataset": ds_name,
                "nb_lignes":   len(df),
                "nb_colonnes": len(df.columns),
                "ram":         _ram_size(df),
            })

        return {
            "status":   "success",
            "fichier":  chemin,
            "chargees": chargees,
            "total":    len(chargees),
            "message":  f"{len(chargees)} feuille(s) chargée(s) depuis {Path(chemin).name}.",
        }

    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_list",
    description="Liste tous les datasets chargés en session avec leurs dimensions, source et taille RAM.",
    parameters={"type": "object", "properties": {}, "required": []},
)
def df_list() -> dict:
    ds = _user_datasets()
    if not ds:
        return {
            "status":   "success",
            "nombre":   0,
            "datasets": [],
            "message":  "Aucun dataset en mémoire. Utilisez df_read pour charger un fichier.",
        }
    datasets = []
    for nom, info in ds.items():
        df = info["df"]
        datasets.append({
            "nom":         nom,
            "source":      info["source"],
            "nb_lignes":   len(df),
            "nb_colonnes": len(df.columns),
            "colonnes":    list(df.columns),
            "ram":         _ram_size(df),
            "charge_le":   info["loaded_at"],
        })
    return {"status": "success", "nombre": len(datasets), "datasets": datasets}


# ══════════════════════════════════════════════════════════════════════════════
# EXPLORATION
# ══════════════════════════════════════════════════════════════════════════════

@tool(
    name="df_head",
    description=(
        "Affiche les premières ou dernières lignes d'un dataset. "
        "Valeur négative de n = dernières lignes."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "n": {"type": "integer", "description": "Nombre de lignes (défaut: 10). Négatif = fin."},
            "colonnes": {
                "type": "array", "items": {"type": "string"},
                "description": "Sous-ensemble de colonnes (défaut: toutes).",
            },
        },
        "required": ["nom"],
    },
)
def df_head(nom: str, n: int = 10, colonnes: Optional[list] = None) -> dict:
    try:
        df = _get_df(nom)
        n = max(-len(df), min(n, _MAX_ROWS_DISPLAY))
        if colonnes:
            manquantes = [c for c in colonnes if c not in df.columns]
            if manquantes:
                return {"status": "error", "error": f"Colonnes introuvables : {manquantes}. Disponibles : {list(df.columns)}"}
            df = df[colonnes]
        view = df.tail(abs(n)) if n < 0 else df.head(n)
        records, _ = _df_to_records(view, max_rows=abs(n))
        return {
            "status": "success", "nom": nom,
            "nb_lignes": len(_get_df(nom)), "affichees": len(records),
            "sens": "fin" if n < 0 else "début", "lignes": records,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_info",
    description=(
        "Statistiques descriptives complètes : types, valeurs manquantes, "
        "min/max/moyenne/médiane/écart-type pour les numériques, "
        "top valeurs pour les catégorielles, plage pour les dates. "
        "Signale les colonnes de dates mal typées."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "colonnes": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes à analyser (défaut: toutes).",
            },
        },
        "required": ["nom"],
    },
)
def df_info(nom: str, colonnes: Optional[list] = None) -> dict:
    try:
        df = _get_df(nom)
        if colonnes:
            manquantes = [c for c in colonnes if c not in df.columns]
            if manquantes:
                return {"status": "error", "error": f"Colonnes introuvables : {manquantes}"}
            df = df[colonnes]

        infos = []
        for col in df.columns:
            series = df[col]
            nb_null = int(series.isna().sum())
            nb_unique = int(series.nunique(dropna=True))
            col_info: dict = {
                "colonne":       col,
                "type":          str(series.dtype),
                "nb_valeurs":    len(series),
                "nb_manquants":  nb_null,
                "pct_manquants": round(nb_null / len(series) * 100, 1) if len(series) else 0,
                "nb_uniques":    nb_unique,
            }
            if pd.api.types.is_numeric_dtype(series):
                desc = series.describe()
                col_info.update({
                    "min": _safe(desc.get("min")), "max": _safe(desc.get("max")),
                    "moyenne": _safe(desc.get("mean")), "mediane": _safe(series.median()),
                    "ecart_type": _safe(desc.get("std")),
                    "q25": _safe(desc.get("25%")), "q75": _safe(desc.get("75%")),
                })
            elif pd.api.types.is_datetime64_any_dtype(series):
                col_info.update({"min": _safe(series.min()), "max": _safe(series.max())})
            else:
                top = series.value_counts(dropna=True).head(5)
                col_info["top_valeurs"] = [
                    {"valeur": _safe(v), "occurrences": int(c)} for v, c in top.items()
                ]
                # Détection date mal typée
                sample = series.dropna().head(20)
                if len(sample) >= 3:
                    try:
                        pd.to_datetime(sample, infer_datetime_format=True, dayfirst=True)
                        col_info["conseil"] = "Colonne potentiellement de type date — utiliser df_cast avec type 'datetime'."
                    except Exception:
                        pass
            infos.append(col_info)

        return {
            "status":       "success",
            "nom":          nom,
            "nb_lignes":    len(df),
            "nb_colonnes":  len(df.columns),
            "nb_doublons":  int(df.duplicated().sum()),
            "ram":          _ram_size(df),
            "colonnes":     infos,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_value_counts",
    description="Fréquence des valeurs d'une colonne. Permet de détecter les valeurs dominantes et la distribution.",
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "colonne": {"type": "string", "description": "Colonne à analyser."},
            "limite": {"type": "integer", "description": "Nombre max de valeurs (défaut: 20)."},
            "normaliser": {"type": "boolean", "description": "Retourner les pourcentages (défaut: false)."},
            "inclure_nan": {"type": "boolean", "description": "Inclure les valeurs manquantes (défaut: false)."},
        },
        "required": ["nom", "colonne"],
    },
)
def df_value_counts(
    nom: str, colonne: str,
    limite: int = 20, normaliser: bool = False, inclure_nan: bool = False,
) -> dict:
    try:
        df = _get_df(nom)
        if colonne not in df.columns:
            return {"status": "error", "error": f"Colonne '{colonne}' introuvable. Disponibles : {list(df.columns)}"}
        vc = df[colonne].value_counts(normalize=normaliser, dropna=not inclure_nan)
        total = len(df[colonne].dropna() if not inclure_nan else df[colonne])
        resultats = []
        for val, count in vc.head(limite).items():
            entry: dict = {"valeur": _safe(val)}
            if normaliser:
                entry["pourcentage"] = round(float(count) * 100, 2)
            else:
                entry["occurrences"] = int(count)
                entry["pourcentage"] = round(int(count) / total * 100, 2) if total else 0
            resultats.append(entry)
        return {
            "status": "success", "nom": nom, "colonne": colonne,
            "nb_valeurs_total": total,
            "nb_valeurs_uniques": int(df[colonne].nunique(dropna=not inclure_nan)),
            "nb_manquants": int(df[colonne].isna().sum()),
            "affichees": len(resultats), "resultats": resultats,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


# ══════════════════════════════════════════════════════════════════════════════
# ANALYSE
# ══════════════════════════════════════════════════════════════════════════════

@tool(
    name="df_groupby",
    description=(
        "Agrège un dataset par groupe (GROUP BY). "
        "Fonctions : sum, mean, count, min, max, median, std, nunique, first, last. "
        "Support des périodes temporelles : grouper une colonne date par 'M' (mois), "
        "'Q' (trimestre), 'Y' (année), 'W' (semaine)."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "grouper_par": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes de regroupement.",
            },
            "agregations": {
                "type": "object",
                "description": "Dictionnaire colonne → fonction(s). Ex: {'ca': 'sum', 'age': 'median'}.",
            },
            "periode": {
                "type": "object",
                "description": (
                    "Optionnel : {'colonne': 'date_entree', 'freq': 'M'} pour grouper par période temporelle. "
                    "freq : 'D' jour, 'W' semaine, 'M' mois, 'Q' trimestre, 'Y' année."
                ),
            },
            "trier_par": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes de tri du résultat.",
            },
            "ordre_desc": {"type": "boolean", "description": "Tri décroissant (défaut: false)."},
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset résultat."},
        },
        "required": ["nom", "grouper_par", "agregations"],
    },
)
def df_groupby(
    nom: str,
    grouper_par: list,
    agregations: dict,
    periode: Optional[dict] = None,
    trier_par: Optional[list] = None,
    ordre_desc: bool = False,
    sauvegarder_sous: Optional[str] = None,
) -> dict:
    AGGS = {"sum", "mean", "count", "min", "max", "median", "std", "nunique", "first", "last"}
    try:
        df = _get_df(nom).copy()

        # Groupby par période temporelle
        if periode:
            col_date = periode.get("colonne")
            freq = periode.get("freq", "M")
            if col_date not in df.columns:
                return {"status": "error", "error": f"Colonne date '{col_date}' introuvable."}
            df[col_date] = pd.to_datetime(df[col_date], errors="coerce", dayfirst=True)
            df[col_date] = df[col_date].dt.to_period(freq).astype(str)
            if col_date not in grouper_par:
                grouper_par = [col_date] + grouper_par

        manquantes = [c for c in grouper_par if c not in df.columns]
        if manquantes:
            return {"status": "error", "error": f"Colonnes de groupement introuvables : {manquantes}"}

        agg_dict: dict = {}
        for col, funcs in agregations.items():
            if col not in df.columns:
                return {"status": "error", "error": f"Colonne '{col}' introuvable."}
            funcs_list = [funcs] if isinstance(funcs, str) else list(funcs)
            invalides = [f for f in funcs_list if f not in AGGS]
            if invalides:
                return {"status": "error", "error": f"Fonctions inconnues : {invalides}. Disponibles : {sorted(AGGS)}"}
            agg_dict[col] = funcs_list if len(funcs_list) > 1 else funcs_list[0]

        result = df.groupby(grouper_par, dropna=False).agg(agg_dict)
        if isinstance(result.columns, pd.MultiIndex):
            result.columns = ["_".join(str(c) for c in col).strip("_") for col in result.columns]
        result = result.reset_index()

        sort_cols = [c for c in (trier_par or grouper_par) if c in result.columns]
        if sort_cols:
            result = result.sort_values(sort_cols, ascending=not ordre_desc)

        if sauvegarder_sous:
            _store(sauvegarder_sous, result.reset_index(drop=True), f"df_groupby({nom})")

        records, truncated = _df_to_records(result)
        return {
            "status": "success", "nom_source": nom,
            "grouper_par": grouper_par, "nb_groupes": len(result),
            "tronque": truncated, "sauvegarde_sous": sauvegarder_sous,
            "colonnes": list(result.columns), "lignes": records,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_resample",
    description=(
        "Agrège un dataset par période temporelle : évolution mensuelle, trimestrielle, annuelle, etc. "
        "Idéal pour analyser des séries temporelles RH (effectifs, absences, recrutements par mois). "
        "Fréquences : 'D' (jour), 'W' (semaine), 'M' (mois), 'Q' (trimestre), 'Y' (année)."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "colonne_date": {"type": "string", "description": "Colonne de dates à utiliser pour le regroupement."},
            "freq": {
                "type": "string",
                "description": "Fréquence de regroupement : 'D', 'W', 'M', 'Q', 'Y' (défaut: 'M').",
            },
            "agregations": {
                "type": "object",
                "description": "Dictionnaire colonne → fonction(s). Ex: {'effectif': 'count', 'salaire': 'mean'}.",
            },
            "colonnes_groupe": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes supplémentaires de regroupement (ex: ['service', 'site']).",
            },
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset résultat."},
        },
        "required": ["nom", "colonne_date", "agregations"],
    },
)
def df_resample(
    nom: str,
    colonne_date: str,
    agregations: dict,
    freq: str = "M",
    colonnes_groupe: Optional[list] = None,
    sauvegarder_sous: Optional[str] = None,
) -> dict:
    FREQS = {"D", "W", "M", "Q", "Y", "QS", "MS", "YS"}
    if freq not in FREQS:
        return {"status": "error", "error": f"Fréquence '{freq}' invalide. Valeurs : {sorted(FREQS)}"}
    try:
        df = _get_df(nom).copy()
        if colonne_date not in df.columns:
            return {"status": "error", "error": f"Colonne date '{colonne_date}' introuvable."}

        df[colonne_date] = pd.to_datetime(df[colonne_date], errors="coerce", dayfirst=True)
        nb_nat = int(df[colonne_date].isna().sum())

        AGGS = {"sum", "mean", "count", "min", "max", "median", "std", "nunique", "first", "last"}
        agg_dict = {}
        for col, funcs in agregations.items():
            if col not in df.columns:
                return {"status": "error", "error": f"Colonne '{col}' introuvable."}
            funcs_list = [funcs] if isinstance(funcs, str) else list(funcs)
            invalides = [f for f in funcs_list if f not in AGGS]
            if invalides:
                return {"status": "error", "error": f"Fonctions inconnues : {invalides}."}
            agg_dict[col] = funcs_list if len(funcs_list) > 1 else funcs_list[0]

        df_valid = df.dropna(subset=[colonne_date])

        if colonnes_groupe:
            manquantes = [c for c in colonnes_groupe if c not in df.columns]
            if manquantes:
                return {"status": "error", "error": f"Colonnes de groupe introuvables : {manquantes}"}
            df_valid[colonne_date] = df_valid[colonne_date].dt.to_period(freq).astype(str)
            result = df_valid.groupby([colonne_date] + colonnes_groupe).agg(agg_dict)
        else:
            df_valid = df_valid.set_index(colonne_date)
            result = df_valid.resample(freq).agg(agg_dict)

        if isinstance(result.columns, pd.MultiIndex):
            result.columns = ["_".join(str(c) for c in col).strip("_") for col in result.columns]
        result = result.reset_index()
        result = result.sort_values(colonne_date)

        if sauvegarder_sous:
            _store(sauvegarder_sous, result.reset_index(drop=True), f"df_resample({nom}, freq={freq})")

        records, truncated = _df_to_records(result, max_rows=_MAX_ROWS_DISPLAY)
        avertissements = []
        if nb_nat > 0:
            avertissements.append(f"{nb_nat} ligne(s) ignorée(s) : date non parseable.")

        return {
            "status": "success", "nom_source": nom, "freq": freq,
            "nb_periodes": len(result), "tronque": truncated,
            "avertissements": avertissements,
            "sauvegarde_sous": sauvegarder_sous,
            "colonnes": list(result.columns), "lignes": records,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_correlate",
    description=(
        "Matrice de corrélation entre colonnes numériques. "
        "Méthodes : 'pearson' (linéaire), 'spearman' (rang, robuste), 'kendall'. "
        "Retourne aussi les paires les plus corrélées."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "colonnes": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes numériques (défaut: toutes).",
            },
            "methode": {"type": "string", "description": "'pearson' (défaut), 'spearman', 'kendall'."},
            "seuil": {"type": "number", "description": "Seuil absolu pour filtrer les paires faibles (défaut: 0.0)."},
            "top_n": {"type": "integer", "description": "Nombre de paires à retourner (défaut: 10)."},
        },
        "required": ["nom"],
    },
)
def df_correlate(
    nom: str, colonnes: Optional[list] = None,
    methode: str = "pearson", seuil: float = 0.0, top_n: int = 10,
) -> dict:
    METHODES = {"pearson", "spearman", "kendall"}
    if methode not in METHODES:
        return {"status": "error", "error": f"Méthode invalide. Disponibles : {sorted(METHODES)}"}
    try:
        df = _get_df(nom)
        if colonnes:
            manquantes = [c for c in colonnes if c not in df.columns]
            if manquantes:
                return {"status": "error", "error": f"Colonnes introuvables : {manquantes}"}
            num_df = df[colonnes].select_dtypes(include="number")
        else:
            num_df = df.select_dtypes(include="number")

        if len(num_df.columns) < 2:
            return {"status": "error", "error": "Au moins 2 colonnes numériques nécessaires."}

        corr = num_df.corr(method=methode)
        matrice = {
            col: {c: round(_safe(v), 4) for c, v in row.items()}
            for col, row in corr.to_dict().items()
        }
        paires = []
        cols = list(corr.columns)
        for i in range(len(cols)):
            for j in range(i + 1, len(cols)):
                val = corr.iloc[i, j]
                if not math.isnan(val) and abs(val) >= seuil:
                    paires.append({
                        "colonne_a": cols[i], "colonne_b": cols[j],
                        "correlation": round(float(val), 4),
                        "intensite": "forte" if abs(val) >= 0.7 else "modérée" if abs(val) >= 0.4 else "faible",
                        "sens": "positive" if val > 0 else "négative",
                    })
        paires.sort(key=lambda x: abs(x["correlation"]), reverse=True)
        return {
            "status": "success", "nom": nom, "methode": methode,
            "nb_colonnes": len(num_df.columns), "colonnes": list(num_df.columns),
            "matrice": matrice, "top_paires": paires[:top_n],
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_outliers",
    description=(
        "Détecte les valeurs aberrantes dans les colonnes numériques. "
        "Méthode IQR (robuste, défaut) ou z-score (distribution normale)."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "colonnes": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes numériques (défaut: toutes).",
            },
            "methode": {"type": "string", "description": "'iqr' (défaut) ou 'zscore'."},
            "k": {"type": "number", "description": "Multiplicateur IQR (défaut: 1.5, extrêmes: 3.0)."},
            "seuil_z": {"type": "number", "description": "Seuil z-score (défaut: 3.0)."},
            "sauvegarder_sous": {"type": "string", "description": "Sauvegarde les lignes aberrantes."},
        },
        "required": ["nom"],
    },
)
def df_outliers(
    nom: str, colonnes: Optional[list] = None,
    methode: str = "iqr", k: float = 1.5, seuil_z: float = 3.0,
    sauvegarder_sous: Optional[str] = None,
) -> dict:
    if methode not in ("iqr", "zscore"):
        return {"status": "error", "error": "Méthode invalide : 'iqr' ou 'zscore'."}
    try:
        df = _get_df(nom)
        num_df = df[colonnes].select_dtypes(include="number") if colonnes else df.select_dtypes(include="number")
        if num_df.empty:
            return {"status": "error", "error": "Aucune colonne numérique trouvée."}

        outlier_mask = pd.Series(False, index=df.index)
        stats_par_colonne = []

        for col in num_df.columns:
            series = num_df[col].dropna()
            if len(series) < 4:
                continue
            if methode == "iqr":
                q1, q3 = series.quantile(0.25), series.quantile(0.75)
                iqr = q3 - q1
                bb, bh = q1 - k * iqr, q3 + k * iqr
                mask_col = (df[col] < bb) | (df[col] > bh)
                stats_par_colonne.append({
                    "colonne": col, "methode": "iqr",
                    "q1": round(float(q1), 4), "q3": round(float(q3), 4),
                    "borne_basse": round(float(bb), 4), "borne_haute": round(float(bh), 4),
                    "nb_outliers": int(mask_col.sum()),
                    "pct_outliers": round(mask_col.sum() / len(df) * 100, 2),
                })
            else:
                z = (df[col] - series.mean()) / series.std()
                mask_col = z.abs() > seuil_z
                stats_par_colonne.append({
                    "colonne": col, "methode": "zscore",
                    "moyenne": round(float(series.mean()), 4),
                    "ecart_type": round(float(series.std()), 4),
                    "seuil_z": seuil_z,
                    "nb_outliers": int(mask_col.sum()),
                    "pct_outliers": round(mask_col.sum() / len(df) * 100, 2),
                })
            outlier_mask = outlier_mask | mask_col.fillna(False)

        outliers_df = df[outlier_mask].copy()
        if sauvegarder_sous:
            _store(sauvegarder_sous, outliers_df.reset_index(drop=True), f"df_outliers({nom})")

        records, truncated = _df_to_records(outliers_df)
        return {
            "status": "success", "nom": nom, "methode": methode,
            "nb_lignes_total": len(df),
            "nb_outliers": int(outlier_mask.sum()),
            "pct_outliers": round(outlier_mask.sum() / len(df) * 100, 2) if len(df) else 0,
            "stats_par_colonne": stats_par_colonne,
            "tronque": truncated, "sauvegarde_sous": sauvegarder_sous, "lignes": records,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_compare",
    description=(
        "Compare deux versions d'un dataset pour détecter les changements : "
        "lignes ajoutées, supprimées, modifiées. "
        "Utile pour auditer les mises à jour d'un fichier RH entre deux exports."
    ),
    parameters={
        "type": "object",
        "properties": {
            "avant": {"type": "string", "description": "Nom du dataset de référence (ancienne version)."},
            "apres": {"type": "string", "description": "Nom du dataset à comparer (nouvelle version)."},
            "cle": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonne(s) clé(s) d'identification des lignes (ex: ['matricule']).",
            },
            "colonnes_comparees": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes à comparer pour détecter les modifications (défaut: toutes).",
            },
            "sauvegarder_ajouts": {"type": "string", "description": "Nom dataset pour les lignes ajoutées."},
            "sauvegarder_suppressions": {"type": "string", "description": "Nom dataset pour les lignes supprimées."},
            "sauvegarder_modifications": {"type": "string", "description": "Nom dataset pour les lignes modifiées."},
        },
        "required": ["avant", "apres", "cle"],
    },
)
def df_compare(
    avant: str, apres: str, cle: list,
    colonnes_comparees: Optional[list] = None,
    sauvegarder_ajouts: Optional[str] = None,
    sauvegarder_suppressions: Optional[str] = None,
    sauvegarder_modifications: Optional[str] = None,
) -> dict:
    try:
        df_av = _get_df(avant)
        df_ap = _get_df(apres)

        for c in cle:
            if c not in df_av.columns:
                return {"status": "error", "error": f"Clé '{c}' absente du dataset '{avant}'."}
            if c not in df_ap.columns:
                return {"status": "error", "error": f"Clé '{c}' absente du dataset '{apres}'."}

        df_av_idx = df_av.set_index(cle)
        df_ap_idx = df_ap.set_index(cle)

        idx_av = set(df_av_idx.index.tolist() if len(cle) > 1 else df_av_idx.index)
        idx_ap = set(df_ap_idx.index.tolist() if len(cle) > 1 else df_ap_idx.index)

        ajouts_idx      = idx_ap - idx_av
        suppressions_idx = idx_av - idx_ap
        communs_idx     = idx_av & idx_ap

        df_ajouts       = df_ap_idx.loc[list(ajouts_idx)].reset_index() if ajouts_idx else pd.DataFrame(columns=df_ap.columns)
        df_suppressions = df_av_idx.loc[list(suppressions_idx)].reset_index() if suppressions_idx else pd.DataFrame(columns=df_av.columns)

        # Détection des modifications sur les lignes communes
        cols_comp = colonnes_comparees or [c for c in df_av_idx.columns if c in df_ap_idx.columns]
        modifications = []
        for idx in communs_idx:
            row_av = df_av_idx.loc[idx, cols_comp] if cols_comp else df_av_idx.loc[idx]
            row_ap = df_ap_idx.loc[idx, cols_comp] if cols_comp else df_ap_idx.loc[idx]
            try:
                changed = ~(row_av.fillna("__NA__").astype(str) == row_ap.fillna("__NA__").astype(str))
                if changed.any():
                    champs = []
                    for col in changed[changed].index:
                        champs.append({
                            "colonne": col,
                            "avant": _safe(row_av[col]),
                            "apres": _safe(row_ap[col]),
                        })
                    cle_val = idx if len(cle) == 1 else dict(zip(cle, idx))
                    modifications.append({"cle": _safe(cle_val), "champs_modifies": champs})
            except Exception:
                pass

        df_modif = pd.DataFrame(modifications) if modifications else pd.DataFrame()

        if sauvegarder_ajouts and not df_ajouts.empty:
            _store(sauvegarder_ajouts, df_ajouts, f"df_compare ajouts ({avant}→{apres})")
        if sauvegarder_suppressions and not df_suppressions.empty:
            _store(sauvegarder_suppressions, df_suppressions, f"df_compare suppressions ({avant}→{apres})")
        if sauvegarder_modifications and not df_modif.empty:
            _store(sauvegarder_modifications, df_modif, f"df_compare modifications ({avant}→{apres})")

        return {
            "status": "success",
            "avant": avant, "apres": apres, "cle": cle,
            "nb_lignes_avant": len(df_av), "nb_lignes_apres": len(df_ap),
            "nb_ajouts": len(df_ajouts), "nb_suppressions": len(df_suppressions),
            "nb_modifications": len(modifications),
            "apercu_ajouts": _df_to_records(df_ajouts, 10)[0],
            "apercu_suppressions": _df_to_records(df_suppressions, 10)[0],
            "apercu_modifications": modifications[:10],
            "message": (
                f"{len(df_ajouts)} ajout(s), {len(df_suppressions)} suppression(s), "
                f"{len(modifications)} modification(s) détectée(s)."
            ),
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


# ══════════════════════════════════════════════════════════════════════════════
# TRANSFORMATION
# ══════════════════════════════════════════════════════════════════════════════

@tool(
    name="df_query",
    description=(
        "Filtre, sélectionne et trie un dataset avec une expression pandas. "
        "Exemples : 'age > 30 and ville == \"Paris\"' — 'salaire.between(30000, 60000)' — "
        "'dept.str.contains(\"Tech\", case=False)'."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset source."},
            "filtre": {"type": "string", "description": "Expression de filtre pandas."},
            "colonnes": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes à conserver (défaut: toutes).",
            },
            "trier_par": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes de tri.",
            },
            "ordre_desc": {"type": "boolean", "description": "Tri décroissant (défaut: false)."},
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset résultat."},
            "limite": {"type": "integer", "description": f"Lignes retournées (défaut: {_MAX_ROWS_DISPLAY})."},
        },
        "required": ["nom"],
    },
)
def df_query(
    nom: str, filtre: Optional[str] = None,
    colonnes: Optional[list] = None, trier_par: Optional[list] = None,
    ordre_desc: bool = False, sauvegarder_sous: Optional[str] = None,
    limite: int = _MAX_ROWS_DISPLAY,
) -> dict:
    limite = min(max(1, limite), _MAX_ROWS_DISPLAY)
    try:
        df = _get_df(nom).copy()
        if filtre:
            try:
                df = df.query(filtre, engine="python")
            except Exception as e:
                return {"status": "error", "error": f"Expression invalide : {e}"}
        if colonnes:
            manquantes = [c for c in colonnes if c not in df.columns]
            if manquantes:
                return {"status": "error", "error": f"Colonnes introuvables : {manquantes}"}
            df = df[colonnes]
        if trier_par:
            manquantes = [c for c in trier_par if c not in df.columns]
            if manquantes:
                return {"status": "error", "error": f"Colonnes de tri introuvables : {manquantes}"}
            df = df.sort_values(trier_par, ascending=not ordre_desc)
        if sauvegarder_sous:
            _store(sauvegarder_sous, df.reset_index(drop=True), f"df_query({nom})")
        records, truncated = _df_to_records(df, limite)
        return {
            "status": "success", "nom_source": nom, "filtre": filtre,
            "nb_lignes": len(df), "tronque": truncated,
            "sauvegarde_sous": sauvegarder_sous,
            "colonnes": list(df.columns), "lignes": records,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_pivot",
    description=(
        "Tableau croisé dynamique. "
        "Agrège des données par groupes (sum, mean, count, min, max, median, std, nunique)."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "index": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes index des lignes.",
            },
            "colonnes": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes à déployer en en-têtes (optionnel).",
            },
            "valeurs": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes numériques à agréger.",
            },
            "agregation": {"type": "string", "description": "Fonction d'agrégation (défaut: 'sum')."},
            "totaux": {"type": "boolean", "description": "Ajouter des totaux (défaut: true)."},
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset résultat."},
        },
        "required": ["nom", "index"],
    },
)
def df_pivot(
    nom: str, index: list,
    colonnes: Optional[list] = None, valeurs: Optional[list] = None,
    agregation: str = "sum", totaux: bool = True,
    sauvegarder_sous: Optional[str] = None,
) -> dict:
    AGGS = {"sum", "mean", "count", "min", "max", "median", "std", "nunique"}
    if agregation not in AGGS:
        return {"status": "error", "error": f"Agrégation '{agregation}' invalide. Valeurs : {sorted(AGGS)}"}
    try:
        df = _get_df(nom)
        all_cols = index + (colonnes or []) + (valeurs or [])
        manquantes = [c for c in all_cols if c not in df.columns]
        if manquantes:
            return {"status": "error", "error": f"Colonnes introuvables : {manquantes}"}

        pivot = pd.pivot_table(
            df, index=index,
            columns=colonnes or None,
            values=valeurs or None,
            aggfunc=agregation,
            margins=totaux, margins_name="Total",
        )
        if isinstance(pivot.columns, pd.MultiIndex):
            pivot.columns = [" / ".join(str(c) for c in col).strip() for col in pivot.columns]
        pivot = pivot.reset_index()

        if sauvegarder_sous:
            _store(sauvegarder_sous, pivot, f"df_pivot({nom})")

        records, truncated = _df_to_records(pivot)
        return {
            "status": "success", "nom_source": nom, "agregation": agregation,
            "nb_lignes": len(pivot), "tronque": truncated,
            "sauvegarde_sous": sauvegarder_sous,
            "colonnes": list(pivot.columns), "lignes": records,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_merge",
    description="Fusionne deux datasets (JOIN). Types : 'inner', 'left', 'right', 'outer'.",
    parameters={
        "type": "object",
        "properties": {
            "gauche": {"type": "string", "description": "Dataset gauche."},
            "droite": {"type": "string", "description": "Dataset droit."},
            "sur": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonne(s) commune(s) pour la jointure.",
            },
            "sur_gauche": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonne(s) de jointure dans le dataset gauche (si noms différents).",
            },
            "sur_droite": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonne(s) de jointure dans le dataset droit (si noms différents).",
            },
            "type_jointure": {"type": "string", "description": "'inner' (défaut), 'left', 'right', 'outer'."},
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset résultat."},
        },
        "required": ["gauche", "droite"],
    },
)
def df_merge(
    gauche: str, droite: str,
    sur: Optional[list] = None, sur_gauche: Optional[list] = None, sur_droite: Optional[list] = None,
    type_jointure: str = "inner", sauvegarder_sous: Optional[str] = None,
) -> dict:
    if type_jointure not in {"inner", "left", "right", "outer"}:
        return {"status": "error", "error": f"Type de jointure invalide : '{type_jointure}'."}
    try:
        merged = pd.merge(
            _get_df(gauche), _get_df(droite),
            on=sur or None, left_on=sur_gauche or None, right_on=sur_droite or None,
            how=type_jointure, suffixes=("_g", "_d"),
        )
        if sauvegarder_sous:
            _store(sauvegarder_sous, merged, f"df_merge({gauche},{droite})")
        records, truncated = _df_to_records(merged)
        return {
            "status": "success", "gauche": gauche, "droite": droite,
            "type_jointure": type_jointure,
            "nb_lignes": len(merged), "tronque": truncated,
            "sauvegarde_sous": sauvegarder_sous,
            "colonnes": list(merged.columns), "lignes": records,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_concat",
    description="Empile plusieurs datasets verticalement (UNION). Utile pour consolider des exports mensuels.",
    parameters={
        "type": "object",
        "properties": {
            "noms": {
                "type": "array", "items": {"type": "string"},
                "description": "Datasets à empiler.",
            },
            "jointure": {"type": "string", "description": "'outer' (défaut) ou 'inner'."},
            "ajouter_colonne_source": {"type": "boolean", "description": "Ajouter colonne '_source' (défaut: false)."},
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset résultat."},
        },
        "required": ["noms", "sauvegarder_sous"],
    },
)
def df_concat(
    noms: list, sauvegarder_sous: str,
    jointure: str = "outer", ajouter_colonne_source: bool = False,
) -> dict:
    if jointure not in ("inner", "outer"):
        return {"status": "error", "error": "jointure doit être 'inner' ou 'outer'."}
    if len(noms) < 2:
        return {"status": "error", "error": "Au moins 2 datasets nécessaires."}
    try:
        frames = []
        for n in noms:
            df_n = _get_df(n)
            if ajouter_colonne_source:
                df_n = df_n.copy()
                df_n["_source"] = n
            frames.append(df_n)
        result = pd.concat(frames, axis=0, join=jointure, ignore_index=True)
        _store(sauvegarder_sous, result, f"df_concat({', '.join(noms)})")
        records, truncated = _df_to_records(result)
        return {
            "status": "success", "sources": noms, "jointure": jointure,
            "sauvegarde_sous": sauvegarder_sous,
            "nb_lignes": len(result), "tronque": truncated,
            "colonnes": list(result.columns), "lignes": records,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_clean",
    description=(
        "Nettoyage en une passe : valeurs manquantes, doublons, espaces, "
        "suppression / renommage de colonnes."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "nan_strategie": {
                "type": "string",
                "description": (
                    "'ignorer' (défaut), 'supprimer_lignes', 'supprimer_colonnes', "
                    "'imputer_moyenne', 'imputer_mediane', 'imputer_mode', 'imputer_valeur'."
                ),
            },
            "nan_seuil_colonnes": {"type": "number", "description": "Seuil NaN pour supprimer une colonne (0.0-1.0, défaut: 0.5)."},
            "nan_valeur": {"type": "string", "description": "Valeur de remplacement pour 'imputer_valeur'."},
            "deduplication": {"type": "boolean", "description": "Supprimer les doublons (défaut: false)."},
            "dedup_colonnes": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes pour la déduplication (défaut: toutes).",
            },
            "strip_strings": {"type": "boolean", "description": "Supprimer les espaces en début/fin (défaut: true)."},
            "supprimer_colonnes": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes à supprimer.",
            },
            "renommer_colonnes": {"type": "object", "description": "Dictionnaire ancien_nom → nouveau_nom."},
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset nettoyé (défaut: écrase la source)."},
        },
        "required": ["nom"],
    },
)
def df_clean(
    nom: str,
    nan_strategie: str = "ignorer",
    nan_seuil_colonnes: float = 0.5,
    nan_valeur: Optional[str] = None,
    deduplication: bool = False,
    dedup_colonnes: Optional[list] = None,
    strip_strings: bool = True,
    supprimer_colonnes: Optional[list] = None,
    renommer_colonnes: Optional[dict] = None,
    sauvegarder_sous: Optional[str] = None,
) -> dict:
    NAN_STRATS = {"ignorer", "supprimer_lignes", "supprimer_colonnes",
                  "imputer_moyenne", "imputer_mediane", "imputer_mode", "imputer_valeur"}
    if nan_strategie not in NAN_STRATS:
        return {"status": "error", "error": f"nan_strategie invalide. Valeurs : {sorted(NAN_STRATS)}"}
    try:
        df = _get_df(nom).copy()
        rapport = []
        nb_lignes_init, nb_cols_init = len(df), len(df.columns)

        if supprimer_colonnes:
            existantes = [c for c in supprimer_colonnes if c in df.columns]
            df = df.drop(columns=existantes)
            rapport.append(f"Colonnes supprimées : {existantes}")

        if renommer_colonnes:
            df = df.rename(columns=renommer_colonnes)
            rapport.append(f"Colonnes renommées : {renommer_colonnes}")

        if strip_strings:
            str_cols = df.select_dtypes(include="object").columns
            for c in str_cols:
                df[c] = df[c].str.strip()
            if len(str_cols):
                rapport.append(f"Espaces supprimés sur {len(str_cols)} colonne(s) texte")

        nb_nan_avant = int(df.isna().sum().sum())
        if nan_strategie == "supprimer_lignes":
            df = df.dropna()
            rapport.append(f"Lignes avec NaN supprimées : {nb_lignes_init - len(df)}")
        elif nan_strategie == "supprimer_colonnes":
            cols_a_suppr = [c for c in df.columns if df[c].isna().mean() > nan_seuil_colonnes]
            df = df.drop(columns=cols_a_suppr)
            rapport.append(f"Colonnes >{nan_seuil_colonnes*100:.0f}% NaN supprimées : {cols_a_suppr}")
        elif nan_strategie == "imputer_moyenne":
            num_cols = df.select_dtypes(include="number").columns
            df[num_cols] = df[num_cols].fillna(df[num_cols].mean())
            rapport.append(f"NaN imputés par la moyenne ({len(num_cols)} col.)")
        elif nan_strategie == "imputer_mediane":
            num_cols = df.select_dtypes(include="number").columns
            df[num_cols] = df[num_cols].fillna(df[num_cols].median())
            rapport.append(f"NaN imputés par la médiane ({len(num_cols)} col.)")
        elif nan_strategie == "imputer_mode":
            for c in df.columns:
                mode = df[c].mode()
                if not mode.empty:
                    df[c] = df[c].fillna(mode[0])
            rapport.append("NaN imputés par le mode")
        elif nan_strategie == "imputer_valeur":
            if nan_valeur is None:
                return {"status": "error", "error": "nan_valeur requis pour 'imputer_valeur'."}
            df = df.fillna(nan_valeur)
            rapport.append(f"NaN remplacés par '{nan_valeur}'")

        if nan_strategie != "ignorer":
            rapport.append(f"NaN : {nb_nan_avant} → {int(df.isna().sum().sum())}")

        if deduplication:
            nb_avant = len(df)
            df = df.drop_duplicates(subset=dedup_colonnes or None, keep="first")
            rapport.append(f"Doublons supprimés : {nb_avant - len(df)}")

        dest = sauvegarder_sous or nom
        _store(dest, df.reset_index(drop=True), f"df_clean({nom})")
        return {
            "status": "success", "nom_source": nom, "sauvegarde_sous": dest,
            "nb_lignes_avant": nb_lignes_init, "nb_lignes_apres": len(df),
            "nb_cols_avant": nb_cols_init, "nb_cols_apres": len(df.columns),
            "nb_nan_restants": int(df.isna().sum().sum()),
            "rapport": rapport,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_cast",
    description=(
        "Convertit le type de colonnes. Types : 'int', 'float', 'str', 'bool', 'datetime', 'category'. "
        "Pour les dates avec format : {'type': 'datetime', 'format': '%d/%m/%Y'}."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "conversions": {
                "type": "object",
                "description": "Dictionnaire colonne → type. Ex: {'age': 'int', 'date': {'type': 'datetime', 'format': '%d/%m/%Y'}}.",
            },
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset résultat (défaut: écrase la source)."},
        },
        "required": ["nom", "conversions"],
    },
)
def df_cast(nom: str, conversions: dict, sauvegarder_sous: Optional[str] = None) -> dict:
    TYPES = {"int", "float", "str", "bool", "datetime", "category"}
    try:
        df = _get_df(nom).copy()
        rapport, avertissements = [], []

        for col, type_cible in conversions.items():
            if col not in df.columns:
                return {"status": "error", "error": f"Colonne introuvable : '{col}'"}
            fmt = None
            if isinstance(type_cible, dict):
                fmt = type_cible.get("format")
                type_cible = type_cible.get("type", "")
            if type_cible not in TYPES:
                return {"status": "error", "error": f"Type '{type_cible}' invalide. Valides : {sorted(TYPES)}"}

            type_avant = str(df[col].dtype)
            try:
                if type_cible == "int":
                    df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
                elif type_cible == "float":
                    df[col] = pd.to_numeric(df[col], errors="coerce")
                elif type_cible == "str":
                    df[col] = df[col].astype(str)
                elif type_cible == "bool":
                    df[col] = df[col].map(
                        lambda x: True if str(x).lower() in ("1", "true", "oui", "yes")
                        else (False if str(x).lower() in ("0", "false", "non", "no") else None)
                    )
                elif type_cible == "datetime":
                    df[col] = pd.to_datetime(df[col], format=fmt, errors="coerce", dayfirst=True)
                elif type_cible == "category":
                    df[col] = df[col].astype("category")

                nb_nan = int(df[col].isna().sum())
                rapport.append({"colonne": col, "avant": type_avant, "apres": str(df[col].dtype), "nb_nan": nb_nan})
                if nb_nan > 0 and type_cible in ("int", "float", "datetime"):
                    avertissements.append(f"'{col}' : {nb_nan} valeur(s) non convertible(s) → NaN")
            except Exception as e:
                return {"status": "error", "error": f"Erreur conversion '{col}' : {e}"}

        dest = sauvegarder_sous or nom
        _store(dest, df, f"df_cast({nom})")
        return {
            "status": "success", "nom_source": nom, "sauvegarde_sous": dest,
            "conversions": rapport, "avertissements": avertissements,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_apply",
    description=(
        "Crée de nouvelles colonnes calculées. "
        "Chaque expression : 'nouvelle_colonne = expression'. "
        "Variables disponibles : colonnes du dataset, today/date_auj, now, "
        "pd, np, cut, qcut, to_datetime, Timestamp, Timedelta. "
        "Exemples RH : 'anciennete_ans = (today - date_entree).dt.days / 365.25' — "
        "'tranche_age = cut(age, bins=[0,25,35,45,55,100], labels=[\"<25\",\"25-34\",\"35-44\",\"45-54\",\"55+\"])'."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "expressions": {
                "type": "array", "items": {"type": "string"},
                "description": "Liste d'expressions 'nouvelle_col = formule'.",
            },
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset résultat (défaut: écrase la source)."},
        },
        "required": ["nom", "expressions"],
    },
)
def df_apply(nom: str, expressions: list, sauvegarder_sous: Optional[str] = None) -> dict:
    try:
        df = _get_df(nom).copy()
        rapport = []

        for expr in expressions:
            match = re.match(r"^\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*=\s*(.+)$", expr.strip(), re.DOTALL)
            if not match:
                return {"status": "error", "error": f"Format invalide : '{expr}'. Attendu : 'nom_colonne = expression'."}

            col_dest = match.group(1).strip()
            formule  = match.group(2).strip()

            context = {col: df[col] for col in df.columns}
            context.update({
                "pd": pd, "np": np,
                "today":       pd.Timestamp.today().normalize(),
                "date_auj":    pd.Timestamp.today().normalize(),
                "now":         pd.Timestamp.now(),
                "Timestamp":   pd.Timestamp,
                "Timedelta":   pd.Timedelta,
                "datetime":    datetime,
                "to_datetime": pd.to_datetime,
                "cut":         pd.cut,
                "qcut":        pd.qcut,
                "isna":        pd.isna,
                "notna":       pd.notna,
            })

            try:
                result = _safe_eval(formule, context)
                df[col_dest] = result
                rapport.append({"colonne_cree": col_dest, "expression": formule, "type": str(df[col_dest].dtype)})
            except Exception as e:
                return {"status": "error", "error": f"Erreur dans '{expr}' : {e}"}

        dest = sauvegarder_sous or nom
        _store(dest, df, f"df_apply({nom})")
        records, truncated = _df_to_records(df)
        return {
            "status": "success", "nom_source": nom, "sauvegarde_sous": dest,
            "colonnes_creees": rapport,
            "nb_lignes": len(df), "tronque": truncated,
            "colonnes": list(df.columns), "lignes": records,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_rename",
    description="Renomme des colonnes d'un dataset.",
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "renommage": {"type": "object", "description": "Dictionnaire ancien_nom → nouveau_nom."},
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset résultat."},
        },
        "required": ["nom", "renommage"],
    },
)
def df_rename(nom: str, renommage: dict, sauvegarder_sous: Optional[str] = None) -> dict:
    try:
        df = _get_df(nom)
        manquantes = [c for c in renommage if c not in df.columns]
        if manquantes:
            return {"status": "error", "error": f"Colonnes introuvables : {manquantes}. Disponibles : {list(df.columns)}"}
        df = df.rename(columns=renommage)
        dest = sauvegarder_sous or nom
        _store(dest, df, f"df_rename({nom})")
        return {"status": "success", "nom_source": nom, "sauvegarde_sous": dest,
                "renommage": renommage, "colonnes": list(df.columns)}
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_duplicates",
    description=(
        "Détecte les doublons dans un dataset. "
        "Mode exact : lignes identiques sur les colonnes spécifiées. "
        "Mode fuzzy : détecte les valeurs proches dans une colonne texte "
        "(noms approchants, matricules mal saisis)."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "colonnes": {
                "type": "array", "items": {"type": "string"},
                "description": "Colonnes pour la détection (défaut: toutes).",
            },
            "mode": {
                "type": "string",
                "description": "'exact' (défaut) ou 'fuzzy' (similarité textuelle sur une colonne).",
            },
            "colonne_fuzzy": {
                "type": "string",
                "description": "Colonne texte pour la détection fuzzy (obligatoire si mode='fuzzy').",
            },
            "seuil_fuzzy": {
                "type": "number",
                "description": "Seuil de similarité 0-1 pour le mode fuzzy (défaut: 0.85).",
            },
            "sauvegarder_sous": {"type": "string", "description": "Sauvegarde les doublons détectés."},
        },
        "required": ["nom"],
    },
)
def df_duplicates(
    nom: str,
    colonnes: Optional[list] = None,
    mode: str = "exact",
    colonne_fuzzy: Optional[str] = None,
    seuil_fuzzy: float = 0.85,
    sauvegarder_sous: Optional[str] = None,
) -> dict:
    try:
        df = _get_df(nom)

        if mode == "exact":
            mask = df.duplicated(subset=colonnes or None, keep=False)
            doublons_df = df[mask].copy()
            doublons_df = doublons_df.sort_values(colonnes or list(df.columns))

            if sauvegarder_sous and not doublons_df.empty:
                _store(sauvegarder_sous, doublons_df.reset_index(drop=True), f"df_duplicates({nom})")

            records, truncated = _df_to_records(doublons_df)
            return {
                "status": "success", "nom": nom, "mode": "exact",
                "nb_lignes_total": len(df),
                "nb_doublons": int(mask.sum()),
                "pct_doublons": round(mask.sum() / len(df) * 100, 2) if len(df) else 0,
                "tronque": truncated, "sauvegarde_sous": sauvegarder_sous,
                "lignes": records,
            }

        elif mode == "fuzzy":
            if not colonne_fuzzy:
                return {"status": "error", "error": "colonne_fuzzy obligatoire pour le mode fuzzy."}
            if colonne_fuzzy not in df.columns:
                return {"status": "error", "error": f"Colonne '{colonne_fuzzy}' introuvable."}

            from difflib import SequenceMatcher

            valeurs = df[colonne_fuzzy].fillna("").astype(str).tolist()
            groupes_similaires = []
            traites = set()

            for i in range(len(valeurs)):
                if i in traites:
                    continue
                groupe = [i]
                for j in range(i + 1, len(valeurs)):
                    if j in traites:
                        continue
                    if not valeurs[i] or not valeurs[j]:
                        continue
                    ratio = SequenceMatcher(None, valeurs[i].lower(), valeurs[j].lower()).ratio()
                    if ratio >= seuil_fuzzy and valeurs[i] != valeurs[j]:
                        groupe.append(j)
                        traites.add(j)
                if len(groupe) > 1:
                    traites.add(i)
                    groupes_similaires.append({
                        "valeurs": [valeurs[k] for k in groupe],
                        "indices": groupe,
                        "similarity_min": round(min(
                            SequenceMatcher(None, valeurs[groupe[0]].lower(), valeurs[k].lower()).ratio()
                            for k in groupe[1:]
                        ), 3),
                    })

            return {
                "status": "success", "nom": nom, "mode": "fuzzy",
                "colonne": colonne_fuzzy, "seuil": seuil_fuzzy,
                "nb_groupes_similaires": len(groupes_similaires),
                "groupes": groupes_similaires[:50],
                "message": f"{len(groupes_similaires)} groupe(s) de valeurs similaires détecté(s).",
            }

        else:
            return {"status": "error", "error": "mode invalide : 'exact' ou 'fuzzy'."}

    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


# ══════════════════════════════════════════════════════════════════════════════
# ÉCHANTILLONNAGE
# ══════════════════════════════════════════════════════════════════════════════

@tool(
    name="df_sample",
    description="Tirage aléatoire ou stratifié. Utile pour explorer un grand dataset.",
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "n": {"type": "integer", "description": "Nombre de lignes (exclusif avec fraction)."},
            "fraction": {"type": "number", "description": "Fraction 0-1 (ex: 0.1 = 10%)."},
            "stratifier_par": {"type": "string", "description": "Colonne pour l'échantillonnage stratifié."},
            "graine": {"type": "integer", "description": "Graine aléatoire pour la reproductibilité."},
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset résultat."},
        },
        "required": ["nom"],
    },
)
def df_sample(
    nom: str, n: Optional[int] = None, fraction: Optional[float] = None,
    stratifier_par: Optional[str] = None, graine: Optional[int] = None,
    sauvegarder_sous: Optional[str] = None,
) -> dict:
    if n is None and fraction is None:
        return {"status": "error", "error": "Spécifier 'n' ou 'fraction'."}
    if n is not None and fraction is not None:
        return {"status": "error", "error": "'n' et 'fraction' sont mutuellement exclusifs."}
    if fraction is not None and not (0 < fraction < 1):
        return {"status": "error", "error": "'fraction' doit être entre 0 et 1 (exclus)."}
    try:
        df = _get_df(nom)
        if stratifier_par:
            if stratifier_par not in df.columns:
                return {"status": "error", "error": f"Colonne '{stratifier_par}' introuvable."}
            frac = fraction if fraction else min(n / len(df), 1.0)
            sample = df.groupby(stratifier_par, group_keys=False).apply(
                lambda x: x.sample(frac=frac, random_state=graine)
            )
            if n is not None:
                sample = sample.head(n)
        else:
            if n is not None:
                n = min(n, len(df))
            sample = df.sample(n=n, frac=fraction, random_state=graine)

        sample = sample.reset_index(drop=True)
        if sauvegarder_sous:
            _store(sauvegarder_sous, sample, f"df_sample({nom})")

        records, truncated = _df_to_records(sample)
        return {
            "status": "success", "nom_source": nom,
            "nb_lignes_source": len(df), "nb_lignes_sample": len(sample),
            "fraction_reelle": round(len(sample) / len(df), 4) if len(df) else 0,
            "sauvegarde_sous": sauvegarder_sous, "tronque": truncated,
            "colonnes": list(sample.columns), "lignes": records,
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


# ══════════════════════════════════════════════════════════════════════════════
# CONFORMITÉ RGPD
# ══════════════════════════════════════════════════════════════════════════════

@tool(
    name="df_anonymize",
    description=(
        "Anonymise ou pseudonymise un dataset (RGPD). "
        "Opérations par colonne : 'supprimer', 'hacher' (SHA-256 + sel), "
        "'pseudonymiser' (identifiant opaque reproductible), 'masquer' (→ '***'), "
        "'generaliser_date' (année ou année+mois), "
        "'generaliser_nombre' (arrondi), 'bruit_gaussien' (bruit aléatoire)."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "operations": {
                "type": "object",
                "description": (
                    "Dictionnaire colonne → opération. "
                    "Ex: {'nom': 'supprimer', 'email': 'hacher', 'salaire': 'masquer', "
                    "'date_naissance': {'op': 'generaliser_date', 'precision': 'annee'}, "
                    "'revenu': {'op': 'bruit_gaussien', 'ecart_type': 0.05}}."
                ),
            },
            "sel": {
                "type": "string",
                "description": "Sel pour hachage/pseudonymisation (aléatoire si absent — non reproductible).",
            },
            "sauvegarder_sous": {"type": "string", "description": "Nom du dataset anonymisé (défaut: <nom>_anon)."},
        },
        "required": ["nom", "operations"],
    },
)
def df_anonymize(
    nom: str, operations: dict,
    sel: Optional[str] = None, sauvegarder_sous: Optional[str] = None,
) -> dict:
    OPS = {"supprimer", "hacher", "pseudonymiser", "masquer",
           "generaliser_date", "generaliser_nombre", "bruit_gaussien"}
    try:
        df = _get_df(nom).copy()
        rapport, avertissements = [], []
        sel_eff = sel or secrets.token_hex(16)
        if not sel:
            avertissements.append("Sel aléatoire généré — pseudonymisation non reproductible entre sessions.")

        for col, config in operations.items():
            if col not in df.columns:
                return {"status": "error", "error": f"Colonne introuvable : '{col}'"}
            op = config if isinstance(config, str) else config.get("op", "")
            params = {} if isinstance(config, str) else {k: v for k, v in config.items() if k != "op"}
            if op not in OPS:
                return {"status": "error", "error": f"Opération '{op}' invalide. Valides : {sorted(OPS)}"}

            if op == "supprimer":
                df = df.drop(columns=[col])
                rapport.append({"colonne": col, "operation": "supprimée"})

            elif op in ("hacher", "pseudonymiser"):
                prefix = "P" if op == "pseudonymiser" else "H"
                def _hash(val, _sel=sel_eff, _p=prefix):
                    if pd.isna(val):
                        return None
                    return f"{_p}_{hashlib.sha256(f'{_sel}:{val}'.encode()).hexdigest()[:12]}"
                df[col] = df[col].apply(_hash)
                rapport.append({"colonne": col, "operation": op})

            elif op == "masquer":
                df[col] = df[col].apply(lambda v: None if pd.isna(v) else "***")
                rapport.append({"colonne": col, "operation": "masquée"})

            elif op == "generaliser_date":
                precision = params.get("precision", "annee")
                dates = pd.to_datetime(df[col], errors="coerce")
                if precision == "mois":
                    df[col] = dates.dt.to_period("M").astype(str).where(dates.notna(), None)
                else:
                    df[col] = dates.dt.year.astype("Int64")
                rapport.append({"colonne": col, "operation": f"date → {precision}"})

            elif op == "generaliser_nombre":
                arrondi = int(params.get("arrondi", 10))
                df[col] = (pd.to_numeric(df[col], errors="coerce") / arrondi).round(0) * arrondi
                rapport.append({"colonne": col, "operation": f"arrondi à {arrondi}"})

            elif op == "bruit_gaussien":
                ecart = float(params.get("ecart_type", 0.05))
                num = pd.to_numeric(df[col], errors="coerce")
                bruit = np.random.normal(0, ecart * num.abs().mean(), size=len(num))
                df[col] = (num + bruit).where(num.notna(), None)
                rapport.append({"colonne": col, "operation": f"bruit gaussien σ={ecart*100:.1f}%"})

        dest = sauvegarder_sous or f"{nom}_anon"
        _store(dest, df.reset_index(drop=True), f"df_anonymize({nom})")
        return {
            "status": "success", "nom_source": nom, "sauvegarde_sous": dest,
            "nb_lignes": len(df), "nb_colonnes": len(df.columns),
            "operations": rapport, "avertissements": avertissements,
            "sel_utilise": "fourni" if sel else "aléatoire",
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


# ══════════════════════════════════════════════════════════════════════════════
# ÉCRITURE
# ══════════════════════════════════════════════════════════════════════════════

@tool(
    name="df_write",
    description=(
        "Exporte un dataset en CSV ou Excel dans le VFS. "
        "Pour Excel : en-têtes en gras, largeurs de colonnes auto-ajustées, freeze pane."
    ),
    parameters={
        "type": "object",
        "properties": {
            "nom": {"type": "string", "description": "Nom du dataset."},
            "destination": {
                "type": "string",
                "description": "Chemin VFS de sortie (.csv ou .xlsx). Défaut: /exports/export_<nom>_<timestamp>.",
            },
            "format": {"type": "string", "description": "'csv' ou 'excel'."},
            "separateur": {"type": "string", "description": "Séparateur CSV (défaut: ',')."},
            "inclure_index": {"type": "boolean", "description": "Inclure l'index pandas (défaut: false)."},
            "feuille": {"type": "string", "description": "Nom de la feuille Excel (défaut: nom du dataset)."},
            "datasets_supplementaires": {
                "type": "array", "items": {"type": "string"},
                "description": "Datasets additionnels comme feuilles Excel supplémentaires.",
            },
        },
        "required": ["nom"],
    },
)
def df_write(
    nom: str,
    destination: Optional[str] = None,
    format: Optional[str] = None,
    separateur: str = ",",
    inclure_index: bool = False,
    feuille: Optional[str] = None,
    datasets_supplementaires: Optional[list] = None,
) -> dict:
    from core.virtual_fs import get_vfs

    try:
        df = _get_df(nom)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        if destination:
            vfs_dest = destination if destination.startswith("/") else f"/{destination}"
            ext = Path(vfs_dest).suffix.lower()
            fmt = format or ("excel" if ext in (".xlsx", ".xls") else "csv")
        else:
            fmt = format or "csv"
            ext = ".xlsx" if fmt == "excel" else ".csv"
            vfs_dest = f"/exports/export_{nom}_{ts}{ext}"

        t0 = time.perf_counter()
        buf = _io.BytesIO()

        if fmt == "csv":
            text = df.to_csv(sep=separateur, index=inclure_index, encoding="utf-8-sig")
            buf.write(text.encode("utf-8-sig"))
            mime = "text/csv"

        elif fmt == "excel":
            sheet_name = (feuille or nom)[:31]
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=inclure_index)
                _format_excel_sheet(writer, sheet_name, df)
                if datasets_supplementaires:
                    for ds_nom in datasets_supplementaires:
                        if ds_nom in _user_datasets():
                            df_extra = _user_datasets()[ds_nom]["df"]
                            sname = ds_nom[:31]
                            df_extra.to_excel(writer, sheet_name=sname, index=inclure_index)
                            _format_excel_sheet(writer, sname, df_extra)
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            return {"status": "error", "error": f"Format inconnu : '{fmt}'."}

        data = buf.getvalue()
        vfs = get_vfs()
        vfs.write_bytes(vfs_dest, data, mime_type=mime, overwrite=True)
        elapsed_ms = round((time.perf_counter() - t0) * 1000, 1)
        taille = len(data)
        taille_str = f"{taille/1024:.1f} Ko" if taille < 1_048_576 else f"{taille/1_048_576:.2f} Mo"

        return {
            "status": "success", "nom": nom, "fichier": vfs_dest,
            "format": fmt, "nb_lignes": len(df), "taille": taille_str, "duree_ms": elapsed_ms,
            "message": f"Exporté dans le VFS : {vfs_dest} ({taille_str}).",
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


def _format_excel_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    """Applique le formatage Excel : en-têtes gras, largeurs auto, freeze pane."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = writer.sheets[sheet_name]

        # En-têtes : gras + fond gris clair
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Largeurs automatiques (max du contenu, plafonné à 50)
        for col_idx, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            header_len = len(str(col))
            try:
                max_content = df[col].dropna().astype(str).str.len().max()
                max_content = int(max_content) if not math.isnan(float(max_content)) else 0
            except Exception:
                max_content = 0
            width = min(max(header_len, max_content) + 4, 50)
            ws.column_dimensions[col_letter].width = width

        # Freeze pane sous la ligne d'en-têtes
        ws.freeze_panes = "A2"
    except Exception:
        pass


@tool(
    name="df_write_excel",
    description=(
        "Export Excel multi-feuilles avec mise en forme complète : "
        "en-têtes gras, couleur de fond configurable, largeurs auto, freeze pane, filtre automatique. "
        "Permet de produire des rapports Excel directement exploitables."
    ),
    parameters={
        "type": "object",
        "properties": {
            "feuilles": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "nom_dataset": {"type": "string"},
                        "nom_feuille": {"type": "string"},
                        "couleur_entete": {"type": "string", "description": "Couleur hex sans # (ex: '4472C4')."},
                    },
                    "required": ["nom_dataset"],
                },
                "description": "Liste de feuilles à inclure.",
            },
            "destination": {
                "type": "string",
                "description": "Chemin VFS du fichier .xlsx. Défaut: /exports/rapport_<timestamp>.xlsx.",
            },
            "inclure_index": {"type": "boolean", "description": "Inclure l'index pandas (défaut: false)."},
        },
        "required": ["feuilles"],
    },
)
def df_write_excel(
    feuilles: list,
    destination: Optional[str] = None,
    inclure_index: bool = False,
) -> dict:
    from core.virtual_fs import get_vfs

    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"status": "error", "error": "openpyxl requis pour df_write_excel."}

    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        vfs_dest = destination or f"/exports/rapport_{ts}.xlsx"
        if not vfs_dest.startswith("/"):
            vfs_dest = f"/{vfs_dest}"

        buf = _io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            for feuille_conf in feuilles:
                ds_nom = feuille_conf.get("nom_dataset", "")
                sheet_name = (feuille_conf.get("nom_feuille") or ds_nom)[:31]
                couleur = feuille_conf.get("couleur_entete", "4472C4").lstrip("#")

                if ds_nom not in _user_datasets():
                    return {"status": "error", "error": f"Dataset '{ds_nom}' introuvable."}
                df = _user_datasets()[ds_nom]["df"]
                df.to_excel(writer, sheet_name=sheet_name, index=inclure_index)

                ws = writer.sheets[sheet_name]
                header_font  = Font(bold=True, color="FFFFFF", size=11)
                header_fill  = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")
                header_align = Alignment(horizontal="center", vertical="center")

                for cell in ws[1]:
                    cell.font  = header_font
                    cell.fill  = header_fill
                    cell.alignment = header_align

                for col_idx, col in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    header_len = len(str(col))
                    try:
                        max_content = df[col].dropna().astype(str).str.len().max()
                        max_content = int(max_content) if not math.isnan(float(max_content)) else 0
                    except Exception:
                        max_content = 0
                    ws.column_dimensions[col_letter].width = min(max(header_len, max_content) + 4, 50)

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

        data = buf.getvalue()
        vfs = get_vfs()
        vfs.write_bytes(vfs_dest, data, mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", overwrite=True)

        taille = len(data)
        taille_str = f"{taille/1024:.1f} Ko" if taille < 1_048_576 else f"{taille/1_048_576:.2f} Mo"

        return {
            "status": "success",
            "fichier": vfs_dest,
            "nb_feuilles": len(feuilles),
            "taille": taille_str,
            "message": f"Fichier Excel ({len(feuilles)} feuille(s)) exporté : {vfs_dest} ({taille_str}).",
        }
    except KeyError as e:
        return {"status": "error", "error": str(e)}
    except Exception as e:
        return {"status": "error", "error": f"Erreur : {e}"}


@tool(
    name="df_drop",
    description="Supprime un ou plusieurs datasets de la mémoire session pour libérer des ressources.",
    parameters={
        "type": "object",
        "properties": {
            "noms": {
                "type": "array", "items": {"type": "string"},
                "description": "Noms des datasets à supprimer.",
            },
        },
        "required": ["noms"],
    },
)
def df_drop(noms: list) -> dict:
    ds = _user_datasets()
    supprimes, introuvables = [], []
    for nom in noms:
        if nom in ds:
            del ds[nom]
            supprimes.append(nom)
        else:
            introuvables.append(nom)
    return {
        "status": "success",
        "supprimes": supprimes, "introuvables": introuvables,
        "restants": list(ds.keys()),
        "message": f"{len(supprimes)} dataset(s) supprimé(s).",
    }
